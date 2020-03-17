import requests 
import openpyxl
import os

from lxml import html
err = 0
name = []
img = []
new_img_name = []

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'FRIENDS'

url = 'http://localhost/रितेश%20घिमिरे.html'

path_Name = '//*[@id="pagelet_timeline_app_collection_100007622592333:2356318349:2"]/ul/li/div/div/div[2]/div/div/div/a'
image = '//*[@id="pagelet_timeline_app_collection_100007622592333:2356318349:2"]/ul/li//img/@src'

response = requests.get(url) 

byte_data = response.content 

source_code = html.fromstring(byte_data)

imageSrc = source_code.xpath(image)
tree_Name = source_code.xpath(path_Name) 


for i in range(len(tree_Name)):
    url = 'D:' + imageSrc[i].split('.')[1] + '.jpg'
    sheet.cell(i+1,1).value = i+1
    sheet.cell(i+1,2).value = tree_Name[i].text_content()
    img.append(url)
    name.append(tree_Name[i].text_content())

for i in range(len(name)):
    new_img_name.append('D:/Facebook/' + name[i] + '.jpg')

for i in range(len(name)):
    print(name[i],new_img_name[i])
    try:
        os.rename(img[i], new_img_name[i])
    except Exception as e:
        print(str(e))
        sheet.cell(err+1,5).value = name[i]
        err += 1
print(err)
wb.save('friends_new1.xlsx')

