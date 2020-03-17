# Python3 code implementing web scraping using lxml 

import requests 
import openpyxl
# import only html class 
from lxml import html 
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'QUOTES'

# url to scrape data from 
url = 'https://www.brainyquote.com/lists/topics/top-10-science-quotes'

# path to particular element 
path = '//*[@id="page-body"]/div/div/div/a/ p'

# get response object 
response = requests.get(url) 

# get byte string 
byte_data = response.content 

# get filtered source code 
source_code = html.fromstring(byte_data) 

# jump to preferred html element 
tree = source_code.xpath(path) 

# print texts in first element in list 
for i in range(len(tree)):
    print(tree[i].text_content())
    sheet.cell(i+1,2).value = tree[i].text_content()
    sheet.cell(i+1,1).value = 10-i

wb.save('quote.xlsx')
# //*[@id="page-body"]