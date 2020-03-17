import requests 
import openpyxl

from lxml import html 

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'FRIENDS'

url = 'http://localhost/facebook-riteshghimire397/friends/friends.html'

path_Name = '/html/body/div/div/div/div[2]/div[2]/div/div[1]'
path_Date = '/html/body/div/div/div/div[2]/div[2]/div/div[2]'

response = requests.get(url) 

byte_data = response.content 

source_code = html.fromstring(byte_data) 

tree_Name = source_code.xpath(path_Name) 
tree_Date = source_code.xpath(path_Date) 

for i in range(len(tree_Name)):
    print(tree_Name[i].text_content(),tree_Date[i].text_content())
    sheet.cell(i+1,1).value = i+1
    sheet.cell(i+1,2).value = tree_Name[i].text_content()
    sheet.cell(i+1,3).value = tree_Date[i].text_content() 

wb.save('friends.xlsx')
