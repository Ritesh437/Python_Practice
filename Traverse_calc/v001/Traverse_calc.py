from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QMessageBox
from PyQt5.QtWidgets import QInputDialog
from PyQt5.QtWidgets import QFileDialog

import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font, PatternFill

import math


t_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
th_border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
bold_grey = Font(bold=True, color='a9a9a9')
big_Text = Font(size=20)
color_red = PatternFill(start_color='ffb0b0',fill_type='solid')
color_green = PatternFill(start_color='b1ffb0',fill_type='solid')

def bg_color_range(sheet, cell_range, pattern):
    start_cell, end_cell = cell_range.split(':')
    start_coord = openpyxl.utils.cell.coordinate_from_string(start_cell)
    start_row = start_coord[1]
    start_col = openpyxl.utils.cell.column_index_from_string(start_coord[0])
    end_coord = openpyxl.utils.cell.coordinate_from_string(end_cell)
    end_row = end_coord[1]
    end_col = openpyxl.utils.cell.column_index_from_string(end_coord[0])

    for row in range(start_row, end_row + 1):
        for col_idx in range(start_col, end_col + 1):
            sheet.cell(row,col_idx).fill = pattern

def style_range(sheet, cell_range, style):

    start_cell, end_cell = cell_range.split(':')
    start_coord = openpyxl.utils.cell.coordinate_from_string(start_cell)
    start_row = start_coord[1]
    start_col = openpyxl.utils.cell.column_index_from_string(start_coord[0])
    end_coord = openpyxl.utils.cell.coordinate_from_string(end_cell)
    end_row = end_coord[1]
    end_col = openpyxl.utils.cell.column_index_from_string(end_coord[0])

    for row in range(start_row, end_row + 1):
        for col_idx in range(start_col, end_col + 1):
            sheet.cell(row,col_idx).font = style

def border_range(sheet, cell_range, border):

    start_cell, end_cell = cell_range.split(':')
    start_coord = openpyxl.utils.cell.coordinate_from_string(start_cell)
    start_row = start_coord[1]
    start_col = openpyxl.utils.cell.column_index_from_string(start_coord[0])
    end_coord = openpyxl.utils.cell.coordinate_from_string(end_cell)
    end_row = end_coord[1]
    end_col = openpyxl.utils.cell.column_index_from_string(end_coord[0])

    for row in range(start_row, end_row + 1):
        for col_idx in range(start_col, end_col + 1):
            sheet.cell(row,col_idx).border = border

def set_col_width(sheet,a,b,size):
    for i in range(a,b+1):
        col_str = openpyxl.utils.cell.get_column_letter(i)
        cd = sheet.column_dimensions[col_str]
        cd.width = size

def createTable(path,n_major):
    print(path,n_major)
    wb = openpyxl.Workbook()
    sheet1 = wb.active
    sheet1.sheet_view.showGridLines = False
    sheet1.title = 'Major'
    style_range(sheet1,f'A1:A{sheet1.max_column}',big_Text)
    sheet1.cell(1,1).value = 'Closed Traverse Calculation'
    sheet1.merge_cells('A1:C1')
    sheet1.cell(1,4).value = 'No of Stations : '
    sheet1.merge_cells('D1:E1')
    sheet1.cell(1,6).value = n_major
    sheet1.cell(2,1).value = 'Station'
    sheet1.merge_cells('A2:A3')
    sheet1.cell(2,2).value = 'Leg'
    sheet1.merge_cells('B2:B3')
    sheet1.cell(2,3).value = 'Leg Length'
    sheet1.merge_cells('C2:C3')
    sheet1.cell(2,4).value = 'Hz Angle'
    sheet1.merge_cells('D2:F2')
    sheet1.cell(3,4).value = 'D'
    sheet1.cell(3,5).value = 'M'
    sheet1.cell(3,6).value = 'S'
    sheet1.cell(2,7).value = 'Correction'
    sheet1.merge_cells('G2:I2')
    sheet1.cell(3,7).value = 'D'
    sheet1.cell(3,8).value = 'M'
    sheet1.cell(3,9).value = 'S'
    sheet1.cell(2,10).value = 'Corrected Angles'
    sheet1.merge_cells('J2:L2')
    sheet1.cell(3,10).value = 'D'
    sheet1.cell(3,11).value = 'M'
    sheet1.cell(3,12).value = 'S'
    sheet1.cell(2,13).value = 'Bearing'
    sheet1.merge_cells('M2:O2')
    sheet1.cell(3,13).value = 'D'
    sheet1.cell(3,14).value = 'M'
    sheet1.cell(3,15).value = 'S'
    sheet1.cell(2,16).value = 'Consecutive Coordinates'
    sheet1.merge_cells('P2:Q2')
    sheet1.cell(3,16).value = 'Dep'
    sheet1.cell(3,17).value = 'Lat'
    sheet1.cell(2,18).value = 'Correction'
    sheet1.merge_cells('R2:S2')
    sheet1.cell(3,18).value = 'Dep'
    sheet1.cell(3,19).value = 'Lat'
    sheet1.cell(2,20).value = 'Corrected Consecutive Coordinates'
    sheet1.merge_cells('T2:U2')
    sheet1.cell(3,20).value = 'Dep'
    sheet1.cell(3,21).value = 'Lat'
    sheet1.cell(2,22).value = 'Independent Coordinates'
    sheet1.merge_cells('V2:W2')
    sheet1.cell(3,22).value = 'Easting'
    sheet1.cell(3,23).value = 'Northing'
    sheet1.cell(2,24).value = 'Adjusted Length'
    sheet1.merge_cells('X2:X3')
    sheet1.cell(2,25).value = 'Adjusted Bearing'
    sheet1.merge_cells('Y2:Y3')

    for i in range (1,n_major+1):
        sheet1.cell(i+3,1).value = f'M{i}'
        if(i==n_major):
            sheet1.cell(i+3,2).value = f'M{i}M{1}'
            continue

        sheet1.cell(i+3,2).value = f'M{i}M{i+1}'

    for i in range(1,sheet1.max_row+1):
        for j in range(1,sheet1.max_column+1):
            sheet1.cell(i,j).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

    style_range(sheet1,'A2:Y3',bold_grey)
    style_range(sheet1,f'A4:B{sheet1.max_row}',bold_grey)

    border_range(sheet1, f'A2:Y{sheet1.max_row}',t_border)
    border_range(sheet1, f'A2:Y3',th_border)

    rd = sheet1.row_dimensions[1]
    rd.height = 55
    rd = sheet1.row_dimensions[2]
    rd.height = 45

    bg_color_range(sheet1,f'C4:F{sheet1.max_row}',color_red)
    bg_color_range(sheet1,f'G4:Y{sheet1.max_row}',color_green)
    bg_color_range(sheet1,'M4:O4',color_red)
    bg_color_range(sheet1,'V4:W4',color_red)

    set_col_width(sheet1,4,15,6)
    set_col_width(sheet1,22,23,12)
    try:
        wb.save(path)
    except Exception as e:
        Ui_MainWindow.failed('Unexpected error occured while saving file','file save failed')
        return 0 
        
    print('Table File Created >> Traverse.xlsx')
    print('Fill The Data in Red Cells, close the file and Continue From Calculate.py')



def read_from_column(sheet, column, start_row, no_of_data):
    data = []
    for i in range(0,no_of_data):
        data.append(sheet.cell(start_row+i,column).value)
    return data

def write_to_column(sheet, column, start_row, no_of_data, data):
    for i in range(0,no_of_data):
        sheet.cell(start_row+i,column).value = data[i]

def print_dms_list(angle):
    for i in range(len(angle.Deg)):
        print(f'{angle.Deg[i]}° {angle.Min[i]}\' {angle.Sec[i]}\"\n')

class AngleDMS:
    Deg = []
    Min = []
    Sec = []

    def __init__(self, Deg, Min, Sec):
        self.Deg = Deg
        self.Min = Min
        self.Sec = Sec

def calculate_bearing(bea,ang):
    sm = dmsAddition(bea,ang)
    sm_degree = convert_to_deg(sm)
    if(sm_degree>=540):
        sm = dmsSubtraction(sm,[540,0,0])
    else:
        if(sm_degree>=180):
            sm = dmsSubtraction(sm,[180,0,0])
        else:
            sm = dmsAddition(sm,[180,0,0])
    return sm

def round_up(x):
    aX = abs(x)
    return int(math.ceil(aX)*(aX/x)) if x != 0 else 0


def convert_to_dms(Degree):
    Deg = math.trunc(Degree)
    minn = (Degree-Deg)*60
    Min = math.trunc(minn)
    Sec = round((minn-Min)*60, 4)
    return [Deg, Min, Sec]

def convert_to_deg(dms):
    Degree = (dms[0])+(dms[1]/60)+(dms[2]/3600)
    return Degree


def dmsDivision(a,n):
    div = [a[0]/n, a[1]/n, a[2]/n]
    b = (div[0]-math.trunc(div[0]))*60
    div[0] = math.trunc(div[0])
    div[1] = div[1] + b
    c = (div[1]-math.trunc(div[1]))*60
    div[1] = math.trunc(div[1])
    div[2] = div[2] + c
    return div

def dmsSubtraction(a,b):
    c = a[0] + a[1]/60 + a[2]/3600 
    d = b[0] + b[1]/60 + b[2]/3600
    dif = c - d
    ab = convert_to_dms(abs(dif))
    x = ab[0]
    y = ab[1]
    z = ab[2]
    if(dif>0):
        return [x,y,z]
    else:
        if(dif<0):
            return [-x,-y,-z]
        else:
            return [0,0,0]

def dmsAddition(a,b):
    Deg = a[0] + b[0]
    Min = a[1] + b[1]
    Sec = a[2] + b[2]
    if (Sec >= 60):
        Sec -= 60
        Min += 1
    if(Min >= 60):
        Min -= 60
        Deg += 1
    return [Deg, Min, Sec]

def angleCorrection(hz_angle, corr, n):
    Deg = []
    corDeg = []
    Min = []
    corMin = []
    Sec = []
    corSec = []
    check = []
    Degree = []
    for i in range(len(hz_angle.Deg)):
        check.append(convert_to_deg([hz_angle.Deg[i],hz_angle.Min[i],hz_angle.Sec[i]]))
        Degree.append(convert_to_deg([hz_angle.Deg[i],hz_angle.Min[i],hz_angle.Sec[i]]))

    for i in range(len(check)):
        for j in range(i):
            if(check[i]>check[j]):
                b = check[i]
                check[i] = check[j]
                check[j] = b
    second = abs(corr[2])
    a = int((second-math.trunc(second))*n)
    corr1 = [corr[0], corr[1], math.trunc(corr[2])]
    corr2 = [corr[0], corr[1], round_up(corr[2])]
    nth_great = check[a-1] if a != 0 else 0
    if (corr[2]-math.trunc(corr[2]) == 0):
        for i in range (0,n):
            ang = [hz_angle.Deg[i], hz_angle.Min[i], hz_angle.Sec[i]]
            summed = dmsAddition(ang,corr) if corr[0]>=0 and corr[1]>=0 and corr[2]>=0 else dmsSubtraction(ang,ch_sn(corr))
            Deg.append(summed[0])
            Min.append(summed[1])
            Sec.append(summed[2])
            corDeg.append(corr[0])
            corMin.append(corr[1])
            corSec.append(corr[2])
        corr_hz_angle = AngleDMS(Deg, Min, Sec)
    else:
        print(corr1,corr2)
        for i in range(0,n):
            ang = [hz_angle.Deg[i], hz_angle.Min[i], hz_angle.Sec[i]]
            if(Degree[i]>=nth_great):
                summed = dmsAddition(ang,corr2) if corr2[0]>=0 and corr2[1]>=0 and corr2[2]>=0 else dmsSubtraction(ang,ch_sn(corr2))
                corDeg.append(corr2[0])
                corMin.append(corr2[1])
                corSec.append(corr2[2])
            else:
                summed = dmsAddition(ang,corr1) if corr1[0]>=0 and corr1[1]>=0 and corr1[2]>=0 else dmsSubtraction(ang,ch_sn(corr1))
                corDeg.append(corr1[0])
                corMin.append(corr1[1])
                corSec.append(corr1[2])
            Deg.append(summed[0])
            Min.append(summed[1])
            Sec.append(summed[2])
        corr_hz_angle = AngleDMS(Deg, Min, Sec)
    corrections = AngleDMS(corDeg, corMin, corSec)
    return corr_hz_angle ,corrections

def ch_sn(a):
    return [-a[0],-a[1],-a[2]]

def dmsAddition_list(hz_angle):
    sum_angle = [0,0,0]
    for i in range (len(hz_angle.Deg)):
        to_sum = [hz_angle.Deg[i], hz_angle.Min[i], hz_angle.Sec[i]]
        sum_angle = dmsAddition(sum_angle,to_sum)
    return sum_angle

def calculate_lat_dep(leg,bear):
    n_major = len(leg)
    lat = []
    dep = []
    for i in range (n_major):
        bea = convert_to_deg([bear.Deg[i],bear.Min[i],bear.Sec[i]])
        bea_rad = bea*math.pi/180
        lat.append(round(leg[i]*math.cos(bea_rad),5))
        dep.append(round(leg[i]*math.sin(bea_rad),5))
    return lat , dep

def abs_sum(lis):
    sm = 0
    for i in range(len(lis)):
        sm += abs(lis[i])
    return sm

def T_correction_bowditch(leg,ld):
    print('bow')
    n_major = len(leg)
    cor_ld = []
    correc = []
    p = sum(leg)
    er = sum(ld)
    for i in range(n_major):
        corr = round(-(leg[i]/p)*er,5)
        correc.append(corr)
        cor_ld.append(ld[i]+corr)
    return cor_ld , correc

def T_correction_transit(ld):
    print('trans')
    n_major = len(ld)
    cor_ld = []
    correc = []
    p = abs_sum(ld)
    er = sum(ld)
    for i in range(n_major):
        corr = round(-(ld[i]/p)*er,5)
        correc.append(corr)
        cor_ld.append(ld[i]+corr)
    return cor_ld , correc

def qbToWcb(E,N,quad):
    quad = abs(quad)
    if(E>0 and N>0):
        dms = convert_to_dms(quad)
    if(E>0 and N<0):
        dms = convert_to_dms(180-quad)
    if(E<0 and N>0):
        dms = convert_to_dms(360-quad)
    if(E<0 and N<0):
        dms = convert_to_dms(180+quad)
    return dms

def calculate_traverse(path,typ):
    try:
        wb = openpyxl.load_workbook(path,read_only=False)
    except Exception as e:
        Ui_MainWindow.failed('Unexpected error occured while opening file\nFile Doesnot Exist','file open failed')
        return 0

    sheet1 = wb['Major']
    n_major = sheet1.cell(1,6).value
    test = 1
    for i in range(4,4+n_major):
        for j in range(3,7):
            if(sheet1.cell(i,j).value==None):
                test = 0
    if(sheet1.cell(4,13).value == None or sheet1.cell(4,14).value == None or sheet1.cell(4,15).value == None or sheet1.cell(4,22).value == None or sheet1.cell(4,23).value == None):
        test = 0
    if(test == 0):
        Ui_MainWindow.failed('Incomplete Data\nfill all the red boxes in excel file first','file read failed')
        return 0


    leg_lengths = read_from_column(sheet1, 3, 4, n_major)
    print(leg_lengths)

    Deg = read_from_column(sheet1, 4, 4, n_major)
    Min = read_from_column(sheet1, 5, 4, n_major)
    Sec = read_from_column(sheet1, 6, 4, n_major)

    bearing = [sheet1.cell(4,13).value, sheet1.cell(4,14).value, sheet1.cell(4,15).value]
    print(bearing)

    easting = sheet1.cell(4,22).value
    northing = sheet1.cell(4,23).value
    print(northing,'  ',easting)

    hz_angle = AngleDMS(Deg, Min, Sec)

    perimeter = 0
    for i in range (len(leg_lengths)):
        perimeter += leg_lengths[i]

    sum_angle = dmsAddition_list(hz_angle)

    th_sum = (n_major-2)*180

    th_sum_dms = [th_sum,0,0]

    a_error = dmsSubtraction(sum_angle,th_sum_dms)

    a_correction = [-a_error[0], -a_error[1], -a_error[2]]

    a_ind_corr = dmsDivision(a_correction,n_major)

    corr_hz_angle , corrections = angleCorrection(hz_angle,a_ind_corr,n_major)

    corr_sum_angle = dmsAddition_list(corr_hz_angle)

    print('Initial Angles:\n')
    print_dms_list(hz_angle)
    print('SUM : ', sum_angle)
    print('Corrected Angles:\n')
    print_dms_list(corr_hz_angle)
    print('SUM : ', corr_sum_angle)


    bearings = AngleDMS([],[],[])
    bearings.Deg.append(bearing[0])
    bearings.Min.append(bearing[1])
    bearings.Sec.append(bearing[2])


    for i in range(1,n_major):
        initial_b = [bearings.Deg[i-1],bearings.Min[i-1],bearings.Sec[i-1]]
        initial_a = [corr_hz_angle.Deg[i],corr_hz_angle.Min[i],corr_hz_angle.Sec[i]]
        bea = calculate_bearing(initial_b,initial_a)
        bearings.Deg.append(bea[0])
        bearings.Min.append(bea[1])
        bearings.Sec.append(bea[2])

    a=[corr_hz_angle.Deg[0],corr_hz_angle.Min[0],corr_hz_angle.Sec[0]]
    b=[bearings.Deg[n_major-1],bearings.Min[n_major-1],bearings.Sec[n_major-1]]
    check_bea = calculate_bearing(a,b)

    print('Bearings :\n')
    print_dms_list(bearings)

    lat, dep = calculate_lat_dep(leg_lengths,bearings)
    print(lat,dep)
    sum_lat = sum(lat)
    sum_dep = sum(dep)
    print(sum_lat,'  ',sum_dep)

    corr_lat , correction_lat = T_correction_bowditch(leg_lengths,lat) if typ=='bow' else T_correction_transit(lat)
    corr_dep , correction_dep = T_correction_bowditch(leg_lengths,dep) if typ=='bow' else T_correction_transit(dep)

    print(corr_lat,corr_dep)
    corr_lat_sum = round(sum(corr_lat),5)
    corr_dep_sum = round(sum(corr_dep),5)

    print(corr_lat_sum,'  ', corr_dep_sum)

    eastings = []
    northings = []
    eastings.append(easting)
    northings.append(northing)

    for i in range(1,n_major+1):
        if(i==n_major):
            check_easting = eastings[i-1]+corr_dep[i-1]
            check_northing = northings[i-1]+corr_lat[i-1]
        eastings.append(eastings[i-1]+corr_dep[i-1])
        northings.append(northings[i-1]+corr_lat[i-1])

    check_east = eastings[n_major-1] + corr_dep[n_major-1]
    check_north = northings[n_major-1] + corr_lat[n_major-1]

    al = []
    cb = []
    for i in range (n_major):
        if (i == n_major-1):
            dE = eastings[1]-eastings[i]
            dN = northings[1]-northings[i]
        else:
            dE = eastings[i+1]-eastings[i]
            dN = northings[i+1]-northings[i]

        alength = math.sqrt(dE**2+dN**2)
        quad = math.atan(dE/dN)
        wcb = qbToWcb(dE,dN,quad*(180/math.pi))
        al.append(alength)
        cb.append(f'{wcb[0]}° {wcb[1]}\' {wcb[2]}\"')

    print('Eastings , Northings :\n')
    for i in range(n_major):
        print(eastings[i],' , ',northings[i])

    print(check_east,' , ',check_north)

    write_to_column(sheet1, 7 , 4,  n_major , corrections.Deg)
    write_to_column(sheet1, 8 , 4 , n_major , corrections.Min)
    write_to_column(sheet1, 9 , 4 , n_major , corrections.Sec)
    write_to_column(sheet1, 10 , 4 , n_major , corr_hz_angle.Deg)
    write_to_column(sheet1, 11 , 4 , n_major , corr_hz_angle.Min)
    write_to_column(sheet1, 12 , 4 , n_major , corr_hz_angle.Sec)
    write_to_column(sheet1, 13 , 4 , n_major , bearings.Deg)
    write_to_column(sheet1, 14 , 4 , n_major , bearings.Min)
    write_to_column(sheet1, 15 , 4 , n_major , bearings.Sec)
    write_to_column(sheet1, 16 , 4 , n_major , dep)
    write_to_column(sheet1, 17 , 4 , n_major , lat)
    write_to_column(sheet1, 18 , 4 , n_major , correction_dep)
    write_to_column(sheet1, 19 , 4 , n_major , correction_lat)
    write_to_column(sheet1, 20 , 4 , n_major , corr_dep)
    write_to_column(sheet1, 21 , 4 , n_major , corr_lat)
    write_to_column(sheet1, 22 , 4 , n_major , eastings)
    write_to_column(sheet1, 23 , 4 , n_major , northings)
    write_to_column(sheet1, 24 , 4 , n_major , al)
    write_to_column(sheet1, 25 , 4 , n_major , cb)

    sheet1.cell(4+n_major,1).value = 'Totals'
    sheet1.merge_cells(f'A{4+n_major}:B{4+n_major}')
    sheet1.cell(4+n_major,3).value = perimeter
    sheet1.cell(4+n_major,4).value = sum_angle[0]
    sheet1.cell(4+n_major,5).value = sum_angle[1]
    sheet1.cell(4+n_major,6).value = sum_angle[2]
    sheet1.cell(4+n_major,10).value = corr_sum_angle[0]
    sheet1.cell(4+n_major,11).value = corr_sum_angle[1]
    sheet1.cell(4+n_major,12).value = corr_sum_angle[2]
    sheet1.cell(4+n_major,13).value = check_bea[0]
    sheet1.cell(4+n_major,14).value = check_bea[1]
    sheet1.cell(4+n_major,15).value = check_bea[2]
    sheet1.cell(4+n_major,16).value = sum_dep
    sheet1.cell(4+n_major,17).value = sum_lat
    sheet1.cell(4+n_major,20).value = corr_dep_sum
    sheet1.cell(4+n_major,21).value = corr_lat_sum
    sheet1.cell(4+n_major,22).value = check_east
    sheet1.cell(4+n_major,23).value = check_north

    for i in range(1,sheet1.max_column+1):
        sheet1.cell(4+n_major,i).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

    try:
        wb.save(path)
    except Exception as e:
        Ui_MainWindow.failed('Unexpected error occured while saving file\nMake sure to close the excel file after adding data','file save failed')
        return 0

    msg = QMessageBox()
    msg.setIcon(QMessageBox.Information)
    msg.setWindowTitle('Saved')
    msg.setText('Calculation Completed and Saved')
    x = msg.exec_()

class Ui_MainWindow(object):

    state = {
        'path' : '',
        'n' : 0,
        'type' : ''
    }

    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(545, 413)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.group1 = QtWidgets.QGroupBox(self.centralwidget)
        self.group1.setGeometry(QtCore.QRect(10, 10, 520, 131))
        self.group1.setObjectName("group1")
        self.label1 = QtWidgets.QLabel(self.group1)
        self.label1.setGeometry(QtCore.QRect(10, 20, 121, 30))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label1.setFont(font)
        self.label1.setObjectName("label1")
        self.spinBox1 = QtWidgets.QSpinBox(self.group1)
        self.spinBox1.setGeometry(QtCore.QRect(150, 20, 50, 30))
        self.spinBox1.setMouseTracking(True)
        self.spinBox1.setObjectName("spinBox1")
        self.line = QtWidgets.QFrame(self.group1)
        self.line.setGeometry(QtCore.QRect(0, 60, 521, 16))
        self.line.setFrameShape(QtWidgets.QFrame.HLine)
        self.line.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line.setObjectName("line")
        self.create_btn = QtWidgets.QPushButton(self.group1)
        self.create_btn.setGeometry(QtCore.QRect(20, 80, 100, 30))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.create_btn.setFont(font)
        self.create_btn.setObjectName("create_btn")
        self.label_2 = QtWidgets.QLabel(self.group1)
        self.label_2.setGeometry(QtCore.QRect(140, 90, 321, 16))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")
        self.groupBox = QtWidgets.QGroupBox(self.centralwidget)
        self.groupBox.setGeometry(QtCore.QRect(10, 150, 520, 210))
        self.groupBox.setObjectName("groupBox")
        self.label_4 = QtWidgets.QLabel(self.groupBox)
        self.label_4.setGeometry(QtCore.QRect(20, 20, 491, 16))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_4.setFont(font)
        self.label_4.setObjectName("label_4")
        self.label = QtWidgets.QLabel(self.groupBox)
        self.label.setGeometry(QtCore.QRect(20, 40, 141, 16))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.calculate_btn = QtWidgets.QPushButton(self.groupBox)
        self.calculate_btn.setGeometry(QtCore.QRect(20, 140, 101, 30))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.calculate_btn.setFont(font)
        self.calculate_btn.setObjectName("calculate_btn")
        self.label_3 = QtWidgets.QLabel(self.groupBox)
        self.label_3.setGeometry(QtCore.QRect(20, 185, 380, 15))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_3.setFont(font)
        self.label_3.setObjectName("label_3")
        self.trans = QtWidgets.QRadioButton(self.groupBox)
        self.trans.setGeometry(QtCore.QRect(20, 110, 100, 20))
        self.trans.setChecked(True)
        self.trans.setObjectName("trans")
        self.bow = QtWidgets.QRadioButton(self.groupBox)
        self.bow.setGeometry(QtCore.QRect(140, 110, 110, 20))
        self.bow.setChecked(False)
        self.bow.setObjectName("bow")
        self.line_2 = QtWidgets.QFrame(self.groupBox)
        self.line_2.setGeometry(QtCore.QRect(0, 60, 521, 16))
        self.line_2.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_2.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_2.setObjectName("line_2")
        self.label_5 = QtWidgets.QLabel(self.groupBox)
        self.label_5.setGeometry(QtCore.QRect(20, 80, 291, 25))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_5.setFont(font)
        self.label_5.setObjectName("label_5")
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 545, 21))
        self.menubar.setObjectName("menubar")
        self.menuAbout = QtWidgets.QMenu(self.menubar)
        self.menuAbout.setObjectName("menuAbout")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        self.actionCreator = QtWidgets.QAction(MainWindow)
        self.actionCreator.setObjectName("actionCreator")
        self.actionApp = QtWidgets.QAction(MainWindow)
        self.actionApp.setObjectName("actionApp")
        self.actionContact = QtWidgets.QAction(MainWindow)
        self.actionContact.setCheckable(False)
        self.actionContact.setObjectName("actionContact")
        self.menuAbout.addAction(self.actionCreator)
        self.menuAbout.addAction(self.actionApp)
        self.menuAbout.addAction(self.actionContact)
        self.menubar.addAction(self.menuAbout.menuAction())


        self.actionCreator.triggered.connect(self.Creator)
        self.actionApp.triggered.connect(self.App)
        self.actionContact.triggered.connect(self.Contact)

        self.create_btn.clicked.connect(self.create)
        self.calculate_btn.clicked.connect(self.calculate)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)


    def Creator(self):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setText('Author : Ritesh Ghimere\nProfession : Student @Thapathali Engineering Campus\n(batch : 2074/75 to 2078/79)')
        msg.setWindowTitle('Creator')
        x = msg.exec_()

    def App(self):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setText('App : Traverse Calculator\nVersion : 1.1.0.0\n\nUse only for Educational Purpose\n\n\nCopyright : RiteshGhimere © 2019')
        msg.setWindowTitle('App')
        x = msg.exec_()       

    def Contact(self):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setText('Email : cool.rkroking@gmail.com\n\nFeel free to give Feedback or recomend any suggestions,problems')
        msg.setWindowTitle('Contact')
        x = msg.exec_()       
    
    def create(self):
        if(self.spinBox1.value()<3):
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setText("Number of Stations must be grater than 2")
            msg.setWindowTitle("Warning")
            x = msg.exec_()
            return 0

        filename = QFileDialog.getSaveFileName(self.centralwidget,'Single File','d:\'','*.xlsx')
        path = filename[0]
        print(path)
        if(path == None):
            return 0
        self.state['path'] = path
        self.state['n'] = self.spinBox1.value()
        createTable(self.state['path'],self.state['n'])

    def failed(message,title):
        msg = QMessageBox()
        msg.setText(message)
        msg.setIcon(QMessageBox.Critical)
        msg.setWindowTitle(title)
        x = msg.exec_()
        return 0

    def calculate(self):
        if(self.trans.isChecked()):
            self.state['type'] = 'trans'
        if(self.bow.isChecked()):
            self.state['type'] = 'bow'
        path = self.state['path']
        typ = self.state['type']
        calculate_traverse(path,typ)


    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Traverse_calc"))
        self.group1.setTitle(_translate("MainWindow", "Create Gales Table"))
        self.label1.setText(_translate("MainWindow", "No of Stations : "))
        self.create_btn.setText(_translate("MainWindow", "Create"))
        self.label_2.setText(_translate("MainWindow", "This will create an excel file with required gales table"))
        self.groupBox.setTitle(_translate("MainWindow", "Calculate"))
        self.label_4.setText(_translate("MainWindow", "open above created file, fill out all required data marked with red colour"))
        self.label.setText(_translate("MainWindow", "and click on calculate "))
        self.calculate_btn.setText(_translate("MainWindow", "Calculate"))
        self.label_3.setText(_translate("MainWindow", "the calculated data will be saved within same file created above"))
        self.trans.setText(_translate("MainWindow", "Transit Method"))
        self.bow.setText(_translate("MainWindow", "Bowditchs Method"))
        self.label_5.setText(_translate("MainWindow", "Correction Type"))
        self.menuAbout.setTitle(_translate("MainWindow", "About"))
        self.actionCreator.setText(_translate("MainWindow", "Creator"))
        self.actionApp.setText(_translate("MainWindow", "App"))
        self.actionContact.setText(_translate("MainWindow", "Contact"))


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
