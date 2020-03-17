from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QMessageBox
from PyQt5.QtWidgets import QInputDialog
from PyQt5.QtWidgets import QFileDialog

import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font, PatternFill

import math

import pygame
from pygame.locals import*


is_calculated = 'F'

t_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
th_border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
bold_grey = Font(bold=True, color='a9a9a9')
big_Text = Font(size=20)
color_red = PatternFill(start_color='ffb0b0',fill_type='solid')
color_green = PatternFill(start_color='b1ffb0',fill_type='solid')

def bg_color_range(sheet, cell_range, pattern):
    start_cell, end_cell = cell_range.split(':')
    start_coord = openpyxl.utils.cell.coordinate_from_string(start_cell)
    start_row = start_coord[1]
    start_col = openpyxl.utils.cell.column_index_from_string(start_coord[0])
    end_coord = openpyxl.utils.cell.coordinate_from_string(end_cell)
    end_row = end_coord[1]
    end_col = openpyxl.utils.cell.column_index_from_string(end_coord[0])

    for row in range(start_row, end_row + 1):
        for col_idx in range(start_col, end_col + 1):
            sheet.cell(row,col_idx).fill = pattern

def style_range(sheet, cell_range, style):

    start_cell, end_cell = cell_range.split(':')
    start_coord = openpyxl.utils.cell.coordinate_from_string(start_cell)
    start_row = start_coord[1]
    start_col = openpyxl.utils.cell.column_index_from_string(start_coord[0])
    end_coord = openpyxl.utils.cell.coordinate_from_string(end_cell)
    end_row = end_coord[1]
    end_col = openpyxl.utils.cell.column_index_from_string(end_coord[0])

    for row in range(start_row, end_row + 1):
        for col_idx in range(start_col, end_col + 1):
            sheet.cell(row,col_idx).font = style

def border_range(sheet, cell_range, border):

    start_cell, end_cell = cell_range.split(':')
    start_coord = openpyxl.utils.cell.coordinate_from_string(start_cell)
    start_row = start_coord[1]
    start_col = openpyxl.utils.cell.column_index_from_string(start_coord[0])
    end_coord = openpyxl.utils.cell.coordinate_from_string(end_cell)
    end_row = end_coord[1]
    end_col = openpyxl.utils.cell.column_index_from_string(end_coord[0])

    for row in range(start_row, end_row + 1):
        for col_idx in range(start_col, end_col + 1):
            sheet.cell(row,col_idx).border = border

def set_col_width(sheet,a,b,size):
    for i in range(a,b+1):
        col_str = openpyxl.utils.cell.get_column_letter(i)
        cd = sheet.column_dimensions[col_str]
        cd.width = size

def createTable(path,n_major):
    # print(path,n_major)
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.sheet_view.showGridLines = False
    sheet.title = 'Major'
    style_range(sheet,f'A1:A{sheet.max_column}',big_Text)
    sheet.cell(1,1).value = 'Closed Traverse Calculation'
    sheet.merge_cells('A1:C1')
    sheet.cell(1,4).value = 'No of Stations : '
    sheet.merge_cells('D1:E1')
    sheet.cell(1,6).value = n_major
    sheet.cell(2,1).value = 'Station'
    sheet.merge_cells('A2:A3')
    sheet.cell(2,2).value = 'Leg'
    sheet.merge_cells('B2:B3')
    sheet.cell(2,3).value = 'Leg Length'
    sheet.merge_cells('C2:C3')
    sheet.cell(2,4).value = 'Hz Angle'
    sheet.merge_cells('D2:F2')
    sheet.cell(3,4).value = 'D'
    sheet.cell(3,5).value = 'M'
    sheet.cell(3,6).value = 'S'
    sheet.cell(2,7).value = 'Correction'
    sheet.merge_cells('G2:I2')
    sheet.cell(3,7).value = 'D'
    sheet.cell(3,8).value = 'M'
    sheet.cell(3,9).value = 'S'
    sheet.cell(2,10).value = 'Corrected Angles'
    sheet.merge_cells('J2:L2')
    sheet.cell(3,10).value = 'D'
    sheet.cell(3,11).value = 'M'
    sheet.cell(3,12).value = 'S'
    sheet.cell(2,13).value = 'Bearing'
    sheet.merge_cells('M2:O2')
    sheet.cell(3,13).value = 'D'
    sheet.cell(3,14).value = 'M'
    sheet.cell(3,15).value = 'S'
    sheet.cell(2,16).value = 'Consecutive Coordinates'
    sheet.merge_cells('P2:Q2')
    sheet.cell(3,16).value = 'Dep'
    sheet.cell(3,17).value = 'Lat'
    sheet.cell(2,18).value = 'Correction'
    sheet.merge_cells('R2:S2')
    sheet.cell(3,18).value = 'Dep'
    sheet.cell(3,19).value = 'Lat'
    sheet.cell(2,20).value = 'Corrected Consecutive Coordinates'
    sheet.merge_cells('T2:U2')
    sheet.cell(3,20).value = 'Dep'
    sheet.cell(3,21).value = 'Lat'
    sheet.cell(2,22).value = 'Independent Coordinates'
    sheet.merge_cells('V2:W2')
    sheet.cell(3,22).value = 'Easting'
    sheet.cell(3,23).value = 'Northing'
    sheet.cell(2,24).value = 'Adjusted Length'
    sheet.merge_cells('X2:X3')
    sheet.cell(2,25).value = 'Adjusted Bearing'
    sheet.merge_cells('Y2:Y3')

    for i in range (1,n_major+1):
        sheet.cell(i+3,1).value = f'M{i}'
        if(i==n_major):
            sheet.cell(i+3,2).value = f'M{i}M{1}'
            continue

        sheet.cell(i+3,2).value = f'M{i}M{i+1}'

    for i in range(1,sheet.max_row+1):
        for j in range(1,sheet.max_column+1):
            sheet.cell(i,j).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

    style_range(sheet,'A2:Y3',bold_grey)
    style_range(sheet,f'A4:B{sheet.max_row}',bold_grey)

    border_range(sheet, f'A2:Y{sheet.max_row}',t_border)
    border_range(sheet, f'A2:Y3',th_border)

    rd = sheet.row_dimensions[1]
    rd.height = 55
    rd = sheet.row_dimensions[2]
    rd.height = 45

    bg_color_range(sheet,f'C4:F{sheet.max_row}',color_red)
    bg_color_range(sheet,f'G4:Y{sheet.max_row}',color_green)
    bg_color_range(sheet,'M4:O4',color_red)
    bg_color_range(sheet,'V4:W4',color_red)

    set_col_width(sheet,4,15,6)
    set_col_width(sheet,22,23,12)
    try:
        wb.save(path)
    except Exception as e:
        Ui_MainWindow.failed('Unexpected error occured while saving file','file save failed')
        return 0 
        
    # print(f'Table File Created >> {path}')
    # print('Fill The Data in Red Cells, close the file and Continue From Calculate.py')



def read_from_column(sheet, column, start_row, no_of_data):
    data = []
    for i in range(0,no_of_data):
        data.append(sheet.cell(start_row+i,column).value)
    return data

def write_to_column(sheet, column, start_row, no_of_data, data):
    for i in range(0,no_of_data):
        sheet.cell(start_row+i,column).value = data[i]

def print_dms_list(angle):
    for i in range(len(angle.Deg)):
        print(f'{angle.Deg[i]}° {angle.Min[i]}\' {angle.Sec[i]}\"\n')

class AngleDMS:
    Deg = []
    Min = []
    Sec = []

    def __init__(self, Deg, Min, Sec):
        self.Deg = Deg
        self.Min = Min
        self.Sec = Sec

def calculate_bearing(bea,ang):
    sm = dmsAddition(bea,ang)
    sm_degree = convert_to_deg(sm)
    if(sm_degree>=540):
        sm = dmsSubtraction(sm,[540,0,0])
    else:
        if(sm_degree>=180):
            sm = dmsSubtraction(sm,[180,0,0])
        else:
            sm = dmsAddition(sm,[180,0,0])
    return sm

def round_up(x):
    aX = abs(x)
    return int(math.ceil(aX)*(aX/x)) if x != 0 else 0


def convert_to_dms(Degree):
    Deg = math.trunc(Degree)
    minn = (Degree-Deg)*60
    Min = math.trunc(minn)
    Sec = round((minn-Min)*60, 4)
    if(Sec >= 60):
        Sec -= 60
        Min += 1
    if(Min >= 60):
        Min -= 60
        Deg += 1
    return [Deg, Min, Sec]

def convert_to_deg(dms):
    Degree = (dms[0])+(dms[1]/60)+(dms[2]/3600)
    return Degree


def dmsDivision(a,n):
    div = [a[0]/n, a[1]/n, a[2]/n]
    b = (div[0]-math.trunc(div[0]))*60
    div[0] = math.trunc(div[0])
    div[1] = div[1] + b
    c = (div[1]-math.trunc(div[1]))*60
    div[1] = math.trunc(div[1])
    div[2] = div[2] + c
    div = convert_to_deg(div)
    div = convert_to_dms(div)
    return div

def dmsSubtraction(a,b):
    c = a[0] + a[1]/60 + a[2]/3600 
    d = b[0] + b[1]/60 + b[2]/3600
    dif = c - d
    ab = convert_to_dms(abs(dif))
    x = ab[0]
    y = ab[1]
    z = ab[2]
    if(dif>0):
        return [x,y,z]
    else:
        if(dif<0):
            return [-x,-y,-z]
        else:
            return [0,0,0]

def dmsAddition(a,b):
    Deg = a[0] + b[0]
    Min = a[1] + b[1]
    Sec = a[2] + b[2]
    if (Sec >= 60):
        Sec -= 60
        Min += 1
    if(Min >= 60):
        Min -= 60
        Deg += 1
    return [Deg, Min, Sec]

def angleCorrection(hz_angle, corr, n):
    Deg = []
    corDeg = []
    Min = []
    corMin = []
    Sec = []
    corSec = []
    check = []
    Degree = []
    for i in range(len(hz_angle.Deg)):
        check.append(convert_to_deg([hz_angle.Deg[i],hz_angle.Min[i],hz_angle.Sec[i]]))
        Degree.append(convert_to_deg([hz_angle.Deg[i],hz_angle.Min[i],hz_angle.Sec[i]]))

    # print(corr)
    for i in range(len(check)):
        for j in range(i):
            if(check[i]>check[j]):
                b = check[i]
                check[i] = check[j]
                check[j] = b
    second = abs(corr[2])
    a = int(round((second-math.trunc(second))*n))
    # print(a)
    corr1 = [corr[0], corr[1], math.trunc(corr[2])]
    corr2 = [corr[0], corr[1], round_up(corr[2])]
    nth_great = check[a-1] if a != 0 else 0
    # print(nth_great)
    if (corr[2]-math.trunc(corr[2]) == 0):
        for i in range (0,n):
            ang = [hz_angle.Deg[i], hz_angle.Min[i], hz_angle.Sec[i]]
            summed = dmsAddition(ang,corr) if corr[0]>=0 and corr[1]>=0 and corr[2]>=0 else dmsSubtraction(ang,ch_sn(corr))
            Deg.append(summed[0])
            Min.append(summed[1])
            Sec.append(summed[2])
            corDeg.append(corr[0])
            corMin.append(corr[1])
            corSec.append(corr[2])
        corr_hz_angle = AngleDMS(Deg, Min, Sec)
    else:
        print(corr1,corr2)
        for i in range(0,n):
            ang = [hz_angle.Deg[i], hz_angle.Min[i], hz_angle.Sec[i]]
            if(Degree[i]>=nth_great):
                summed = dmsAddition(ang,corr2) if corr2[0]>=0 and corr2[1]>=0 and corr2[2]>=0 else dmsSubtraction(ang,ch_sn(corr2))
                corDeg.append(corr2[0])
                corMin.append(corr2[1])
                corSec.append(corr2[2])
            else:
                summed = dmsAddition(ang,corr1) if corr1[0]>=0 and corr1[1]>=0 and corr1[2]>=0 else dmsSubtraction(ang,ch_sn(corr1))
                corDeg.append(corr1[0])
                corMin.append(corr1[1])
                corSec.append(corr1[2])
            Deg.append(summed[0])
            Min.append(summed[1])
            Sec.append(summed[2])
        corr_hz_angle = AngleDMS(Deg, Min, Sec)
    corrections = AngleDMS(corDeg, corMin, corSec)
    return corr_hz_angle ,corrections

def ch_sn(a):
    return [-a[0],-a[1],-a[2]]

def dmsAddition_list(hz_angle):
    sum_angle = [0,0,0]
    for i in range (len(hz_angle.Deg)):
        to_sum = [hz_angle.Deg[i], hz_angle.Min[i], hz_angle.Sec[i]]
        sum_angle = dmsAddition(sum_angle,to_sum)
    return sum_angle

def calculate_lat_dep(leg,bear):
    n_major = len(leg)
    lat = []
    dep = []
    for i in range (n_major):
        bea = convert_to_deg([bear.Deg[i],bear.Min[i],bear.Sec[i]])
        bea_rad = bea*math.pi/180
        lat.append(leg[i]*math.cos(bea_rad))
        dep.append(leg[i]*math.sin(bea_rad))
    return lat , dep

def abs_sum(lis):
    sm = 0
    for i in range(len(lis)):
        sm += abs(lis[i])
    return sm

def T_correction_bowditch(leg,ld,er):
    # print('bow')
    n_major = len(leg)
    cor_ld = []
    correc = []
    p = sum(leg)
    for i in range(n_major):
        corr = -(leg[i]/p)*er
        correc.append(corr)
        cor_ld.append(ld[i]+corr)
    return cor_ld , correc

def T_correction_transit(ld,er):
    # print('trans')
    n_major = len(ld)
    cor_ld = []
    correc = []
    p = abs_sum(ld)
    # print(p)
    for i in range(n_major):
        corr = -(abs(ld[i]) /p)*er
        correc.append(corr)
        cor_ld.append(ld[i]+corr)
    return cor_ld , correc

def qbToWcb(E,N,quad):
    quad = abs(quad)
    if(E>0 and N>0):
        dms = convert_to_dms(quad)
    if(E>0 and N<0):
        dms = convert_to_dms(180-quad)
    if(E<0 and N>0):
        dms = convert_to_dms(360-quad)
    if(E<0 and N<0):
        dms = convert_to_dms(180+quad)
    return dms

def calculate_traverse(path,typ):
    try:
        wb = openpyxl.load_workbook(path,read_only=False)
    except Exception as e:
        Ui_MainWindow.failed('Unexpected error occured while opening file\nFile Doesnot Exist','file open failed')
        return 0

    sheet = wb['Major']
    n_major = sheet.cell(1,6).value
    test = 1
    for i in range(4,4+n_major):
        for j in range(3,7):
            if(sheet.cell(i,j).value==None):
                test = 0
    if(sheet.cell(4,13).value == None or sheet.cell(4,14).value == None or sheet.cell(4,15).value == None or sheet.cell(4,22).value == None or sheet.cell(4,23).value == None):
        test = 0
    if(test == 0):
        Ui_MainWindow.failed('Incomplete Data\nfill all the red boxes in excel file first','file read failed')
        return 0


    leg_lengths = read_from_column(sheet, 3, 4, n_major)
    # print(leg_lengths)

    Deg = read_from_column(sheet, 4, 4, n_major)
    Min = read_from_column(sheet, 5, 4, n_major)
    Sec = read_from_column(sheet, 6, 4, n_major)

    bearing = [sheet.cell(4,13).value, sheet.cell(4,14).value, sheet.cell(4,15).value]
    # print(bearing)

    easting = sheet.cell(4,22).value
    northing = sheet.cell(4,23).value
    # print(northing,'  ',easting)

    hz_angle = AngleDMS(Deg, Min, Sec)

    perimeter = 0
    for i in range (len(leg_lengths)):
        perimeter += leg_lengths[i]

    sum_angle = dmsAddition_list(hz_angle)

    th_sum = (n_major-2)*180

    th_sum_dms = [th_sum,0,0]

    a_error = dmsSubtraction(sum_angle,th_sum_dms)

    a_correction = [-a_error[0], -a_error[1], -a_error[2]]

    a_ind_corr = dmsDivision(a_correction,n_major)

    corr_hz_angle , corrections = angleCorrection(hz_angle,a_ind_corr,n_major)

    corr_sum_angle = dmsAddition_list(corr_hz_angle)

    # print('Initial Angles:\n')
    # print_dms_list(hz_angle)
    # print('SUM : ', sum_angle)
    # print('Corrected Angles:\n')
    # print_dms_list(corr_hz_angle)
    # print('SUM : ', corr_sum_angle)


    bearings = AngleDMS([],[],[])
    bearings.Deg.append(bearing[0])
    bearings.Min.append(bearing[1])
    bearings.Sec.append(bearing[2])


    for i in range(1,n_major):
        initial_b = [bearings.Deg[i-1],bearings.Min[i-1],bearings.Sec[i-1]]
        initial_a = [corr_hz_angle.Deg[i],corr_hz_angle.Min[i],corr_hz_angle.Sec[i]]
        bea = calculate_bearing(initial_b,initial_a)
        bearings.Deg.append(bea[0])
        bearings.Min.append(bea[1])
        bearings.Sec.append(bea[2])

    a=[corr_hz_angle.Deg[0],corr_hz_angle.Min[0],corr_hz_angle.Sec[0]]
    b=[bearings.Deg[n_major-1],bearings.Min[n_major-1],bearings.Sec[n_major-1]]
    check_bea = calculate_bearing(a,b)

    # print('Bearings :\n')
    # print_dms_list(bearings)

    lat, dep = calculate_lat_dep(leg_lengths,bearings)
    # print(lat,dep)
    sum_lat = sum(lat)
    sum_dep = sum(dep)
    # print(sum_lat,'  ',sum_dep)

    corr_lat , correction_lat = T_correction_bowditch(leg_lengths,lat,sum_lat) if typ=='bow' else T_correction_transit(lat,sum_lat)
    corr_dep , correction_dep = T_correction_bowditch(leg_lengths,dep,sum_dep) if typ=='bow' else T_correction_transit(dep,sum_dep)

    # print(corr_lat,corr_dep)
    corr_lat_sum = round(sum(corr_lat),5)
    corr_dep_sum = round(sum(corr_dep),5)

    # print(corr_lat_sum,'  ', corr_dep_sum)

    eastings = []
    northings = []
    eastings.append(easting)
    northings.append(northing)

    for i in range(1,n_major+1):
        if(i==n_major):
            check_easting = eastings[i-1]+corr_dep[i-1]
            check_northing = northings[i-1]+corr_lat[i-1]
        eastings.append(eastings[i-1]+corr_dep[i-1])
        northings.append(northings[i-1]+corr_lat[i-1])

    check_east = eastings[n_major-1] + corr_dep[n_major-1]
    check_north = northings[n_major-1] + corr_lat[n_major-1]

    print('Eastings , Northings :\n')

    for i in range(n_major):
        print(eastings[i],' , ',northings[i])

    # print(check_east,' , ',check_north)

    al = []
    cb = []
    for i in range (n_major):
        dE = eastings[i+1]-eastings[i]
        dN = northings[i+1]-northings[i]
        print(dE,dN)
        alength = math.sqrt(dE**2+dN**2)
        quad = math.atan(dE/dN)
        wcb = qbToWcb(dE,dN,quad*(180/math.pi))
        al.append(alength)
        cb.append(f'{wcb[0]}° {wcb[1]}\' {wcb[2]}\"')




    write_to_column(sheet, 7 , 4,  n_major , corrections.Deg)
    write_to_column(sheet, 8 , 4 , n_major , corrections.Min)
    write_to_column(sheet, 9 , 4 , n_major , corrections.Sec)
    write_to_column(sheet, 10 , 4 , n_major , corr_hz_angle.Deg)
    write_to_column(sheet, 11 , 4 , n_major , corr_hz_angle.Min)
    write_to_column(sheet, 12 , 4 , n_major , corr_hz_angle.Sec)
    write_to_column(sheet, 13 , 4 , n_major , bearings.Deg)
    write_to_column(sheet, 14 , 4 , n_major , bearings.Min)
    write_to_column(sheet, 15 , 4 , n_major , bearings.Sec)
    write_to_column(sheet, 16 , 4 , n_major , dep)
    write_to_column(sheet, 17 , 4 , n_major , lat)
    write_to_column(sheet, 18 , 4 , n_major , correction_dep)
    write_to_column(sheet, 19 , 4 , n_major , correction_lat)
    write_to_column(sheet, 20 , 4 , n_major , corr_dep)
    write_to_column(sheet, 21 , 4 , n_major , corr_lat)
    write_to_column(sheet, 22 , 4 , n_major , eastings)
    write_to_column(sheet, 23 , 4 , n_major , northings)
    write_to_column(sheet, 24 , 4 , n_major , al)
    write_to_column(sheet, 25 , 4 , n_major , cb)

    sheet.cell(4+n_major,1).value = 'Totals'
    sheet.merge_cells(f'A{4+n_major}:B{4+n_major}')
    sheet.cell(4+n_major,3).value = perimeter
    sheet.cell(4+n_major,4).value = sum_angle[0]
    sheet.cell(4+n_major,5).value = sum_angle[1]
    sheet.cell(4+n_major,6).value = sum_angle[2]
    sheet.cell(4+n_major,10).value = corr_sum_angle[0]
    sheet.cell(4+n_major,11).value = corr_sum_angle[1]
    sheet.cell(4+n_major,12).value = corr_sum_angle[2]
    sheet.cell(4+n_major,13).value = check_bea[0]
    sheet.cell(4+n_major,14).value = check_bea[1]
    sheet.cell(4+n_major,15).value = check_bea[2]
    sheet.cell(4+n_major,16).value = sum_dep
    sheet.cell(4+n_major,17).value = sum_lat
    sheet.cell(4+n_major,20).value = corr_dep_sum
    sheet.cell(4+n_major,21).value = corr_lat_sum
    sheet.cell(4+n_major,22).value = check_east
    sheet.cell(4+n_major,23).value = check_north

    for i in range(1,sheet.max_column+1):
        sheet.cell(4+n_major,i).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

    try:
        wb.save(path)
    except Exception as e:
        Ui_MainWindow.failed('Unexpected error occured while saving file\nMake sure to close the excel file after adding data','file save failed')
        return 0

    print(eastings,northings)
    Ui_MainWindow.state['is_calculated'] = 'T'
    Ui_MainWindow.state['north'] = northings
    Ui_MainWindow.state['east'] = eastings
    Ui_MainWindow.state['ang_err'] = a_error
    Ui_MainWindow.state['lat_err'] = sum_lat
    Ui_MainWindow.state['dep_err'] = sum_dep
    Ui_MainWindow.state['perimeter'] = perimeter


    msg = QMessageBox()
    msg.setIcon(QMessageBox.Information)
    msg.setWindowTitle('Saved')
    msg.setText('Calculation Completed and Saved')
    x = msg.exec_()


def createMTable(path,n):
    n_minor = n
    # print(path,n_minor)
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.sheet_view.showGridLines = False
    sheet.title = 'Minor'

    style_range(sheet,f'A1:A{sheet.max_column}',big_Text)

    sheet.cell(1,1).value = 'Minor Traverse Calculation'
    sheet.merge_cells('A1:C1')
    sheet.cell(1,4).value = 'No of Stations : '
    sheet.merge_cells('D1:E1')
    sheet.cell(1,6).value = n_minor
    sheet.cell(2,1).value = 'Bearing of leg PQ:'
    sheet.merge_cells('A2:B2')
    sheet.cell(3,1).value = 'Bearing of leg RS:'
    sheet.merge_cells('A3:B3')
    sheet.cell(4,1).value = 'Easting of Station p:'
    sheet.merge_cells('A4:B4')
    sheet.cell(5,1).value = 'Easting of Station R:'
    sheet.merge_cells('A5:B5')
    sheet.cell(4,4).value = 'Northing of Station P:'
    sheet.merge_cells('D4:G4')
    sheet.cell(5,4).value = 'Northing of Station R:'
    sheet.merge_cells('D5:G5')
    sheet.cell(6,1).value = 'Station'
    sheet.merge_cells('A6:A7')
    sheet.cell(6,2).value = 'Leg'
    sheet.merge_cells('B6:B7')
    sheet.cell(6,3).value = 'Leg Length'
    sheet.merge_cells('C6:C7')
    sheet.cell(6,4).value = 'Hz Angle'
    sheet.merge_cells('D6:F6')
    sheet.cell(7,4).value = 'D'
    sheet.cell(7,5).value = 'M'
    sheet.cell(7,6).value = 'S'
    sheet.cell(6,7).value = 'Bearing'
    sheet.merge_cells('G6:I6')
    sheet.cell(7,7).value = 'D'
    sheet.cell(7,8).value = 'M'
    sheet.cell(7,9).value = 'S'
    sheet.cell(6,10).value = 'Correction'
    sheet.merge_cells('J6:L6')
    sheet.cell(7,10).value = 'D'
    sheet.cell(7,11).value = 'M'
    sheet.cell(7,12).value = 'S'
    sheet.cell(6,13).value = 'Corrected Bearing'
    sheet.merge_cells('M6:O6')
    sheet.cell(7,13).value = 'D'
    sheet.cell(7,14).value = 'M'
    sheet.cell(7,15).value = 'S'
    sheet.cell(6,16).value = 'Consecutive Coordinates'
    sheet.merge_cells('P6:Q6')
    sheet.cell(7,16).value = 'Dep'
    sheet.cell(7,17).value = 'Lat'
    sheet.cell(6,18).value = 'Correction'
    sheet.merge_cells('R6:S6')
    sheet.cell(7,18).value = 'Dep'
    sheet.cell(7,19).value = 'Lat'
    sheet.cell(6,20).value = 'Corrected Consecutive Coordinates'
    sheet.merge_cells('T6:U6')
    sheet.cell(7,20).value = 'Dep'
    sheet.cell(7,21).value = 'Lat'
    sheet.cell(6,22).value = 'Independent Coordinates'
    sheet.merge_cells('V6:W6')
    sheet.cell(7,22).value = 'Easting'
    sheet.cell(7,23).value = 'Northing'
    sheet.cell(6,24).value = 'Adjusted Length'
    sheet.merge_cells('X6:X7')
    sheet.cell(6,25).value = 'Adjusted Bearing'
    sheet.merge_cells('Y6:Y7')

    sheet.cell(8,1).value = 'P'
    sheet.cell(8,2).value = 'PQ'


    for i in range(1,n_minor+2):
        sheet.cell(i+8,1).value = f'm{i}'
        if(i == 1):
            sheet.cell(i+7,2).value = f'Pm{i}'
            continue
        if(i == n_minor+1):
            sheet.cell(i+7,2).value = f'm{i-1}R'
            continue

        sheet.cell(i+7,2).value = f'm{i-1}m{i}'


    sheet.cell(9+n_minor,1).value = f'R'
    sheet.cell(9+n_minor,2).value = f'RS'

    set_col_width(sheet,4,15,6)
    set_col_width(sheet,22,23,12)


    bg_color_range(sheet,f'C8:C{sheet.max_row-1}',color_red)
    bg_color_range(sheet,f'D8:F{sheet.max_row}',color_red)
    bg_color_range(sheet,f'G8:Y{sheet.max_row}',color_green)
    bg_color_range(sheet,'C4:C5',color_red)
    bg_color_range(sheet,'C2:E3',color_red)
    bg_color_range(sheet,'H4:H5',color_red)

    rd = sheet.row_dimensions[1]
    rd.height = 55
    rd = sheet.row_dimensions[6]
    rd.height = 45

    border_range(sheet, f'A6:Y{sheet.max_row}',t_border)
    border_range(sheet, f'A6:Y7',th_border)

    style_range(sheet,'A6:Y7',bold_grey)
    style_range(sheet,f'A8:B{sheet.max_row}',bold_grey)

    for i in range(1,sheet.max_row+1):
        for j in range(1,sheet.max_column+1):
            sheet.cell(i,j).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

    try:
        wb.save(path)
    except Exception as e:
        Ui_MainWindow.failed('Unexpected error occured while saving file\nMake sure to close the excel file after adding data','file save failed')
        return 0

def bearing_correction(bearings,cor,n):
    corr_bearings = AngleDMS([],[],[])
    correc = AngleDMS([],[],[])
    for i in range(0,n):
        n_cor = dmsDivision(cor,1/(i+1))
        n_cor = [n_cor[0],n_cor[1],round(n_cor[2])]
        bea = [bearings.Deg[i],bearings.Min[i],bearings.Sec[i]]
        summed = dmsAddition(bea,n_cor) if n_cor[0]>=0 and n_cor[1]>=0 and n_cor[2]>=0 else dmsSubtraction(bea,ch_sn(n_cor))
        corr_bearings.Deg.append(summed[0])
        corr_bearings.Min.append(summed[1])
        corr_bearings.Sec.append(summed[2])
        correc.Deg.append(n_cor[0])
        correc.Min.append(n_cor[1])
        correc.Sec.append(n_cor[2])
    return corr_bearings, correc

def calculate_Mtraverse(path,typ):
    # print(path,typ)
    try:
        wb = openpyxl.load_workbook(path,read_only=False)
    except Exception as e:
        Ui_MainWindow.failed('Unexpected error occured while opening file\nFile Doesnot Exist','file open failed')
        return 0

    sheet = wb['Minor']
    n_minor = sheet.cell(1,6).value

    test = 1
    for i in range(8,9+n_minor):
        if(sheet.cell(i,3).value==None):
            test = 0
    for i in range(8,10+n_minor):
        for j in range(4,7):
            if(sheet.cell(i,j).value == None):
                test = 0

    for i in range(2,4):
        for j in range(3,6):
            if(sheet.cell(i,j).value == None):
                test = 0

    for i in range(4,6):
        if(sheet.cell(i,3).value==None):
            test = 0
    
    for i in range(4,6):
        if(sheet.cell(i,8).value==None):
            test = 0

    if(test == 0):
        Ui_MainWindow.failed('Incomplete Data\nfill all the red boxes in excel file first','file read failed')
        return 0

    # print('Complete')

    leg_lengths = read_from_column(sheet, 3, 8, n_minor+1)
    print(leg_lengths)
    perimeter = sum(leg_lengths)
    Deg = read_from_column(sheet, 4, 8, n_minor+2)
    Min = read_from_column(sheet, 5, 8, n_minor+2)
    Sec = read_from_column(sheet, 6, 8, n_minor+2)

    hz_angle = AngleDMS(Deg, Min, Sec)

    initial_bearing = [sheet.cell(2,3).value, sheet.cell(2,4).value, sheet.cell(2,5).value]
    final_bearing = [sheet.cell(3,3).value, sheet.cell(3,4).value, sheet.cell(3,5).value]
    initial_easting = sheet.cell(4,3).value
    final_easting = sheet.cell(5,3).value
    initial_northing = sheet.cell(4,8).value
    final_northing = sheet.cell(5,8).value


    bearings = AngleDMS([], [], [])
    # print(initial_easting,final_easting,initial_northing,final_northing)

    one_bea = calculate_bearing(initial_bearing,[hz_angle.Deg[0],hz_angle.Min[0],hz_angle.Sec[0]])
    bearings.Deg.append(one_bea[0])
    bearings.Min.append(one_bea[1])
    bearings.Sec.append(one_bea[2])
    # print(initial_bearing,one_bea,[hz_angle.Deg[0],hz_angle.Min[0],hz_angle.Sec[0]])
    for i in range(n_minor+1):
        bea = calculate_bearing([bearings.Deg[i],bearings.Min[i],bearings.Sec[i]],[hz_angle.Deg[i+1],hz_angle.Min[i+1],hz_angle.Sec[i+1]])
        bearings.Deg.append(bea[0])
        bearings.Min.append(bea[1])
        bearings.Sec.append(bea[2])

    # print_dms_list(bearings)


    l = len(bearings.Deg)-1

    # print(l)
        
    bea_err = dmsSubtraction([bearings.Deg[l],bearings.Min[l],bearings.Sec[l]],final_bearing)
    # print(bea_err)
    corr = [-bea_err[0],-bea_err[1],-bea_err[2]]
    # print(corr)
    ind_corr = dmsDivision(corr,n_minor+2)
    # print(ind_corr)

    corr_bearings, correc = bearing_correction(bearings,ind_corr,n_minor+2)

    # print_dms_list(corr_bearings)

    write_to_column(sheet, 7, 8, n_minor+2, bearings.Deg)
    write_to_column(sheet, 8, 8, n_minor+2, bearings.Min)
    write_to_column(sheet, 9, 8, n_minor+2, bearings.Sec)
    write_to_column(sheet, 10, 8, n_minor+2, correc.Deg)
    write_to_column(sheet, 11, 8, n_minor+2, correc.Min)
    write_to_column(sheet, 12, 8, n_minor+2, correc.Sec)
    write_to_column(sheet, 13, 8, n_minor+2, corr_bearings.Deg)
    write_to_column(sheet, 14, 8, n_minor+2, corr_bearings.Min)
    write_to_column(sheet, 15, 8, n_minor+2, corr_bearings.Sec)


    lat , dep = calculate_lat_dep(leg_lengths,corr_bearings)
    # print(lat,dep)

    write_to_column(sheet, 16, 8, len(dep),dep)
    write_to_column(sheet, 17, 8, len(lat),lat)

    # print(sum(lat),sum(dep))
    er_lat = sum(lat)-(final_northing-initial_northing)
    er_dep = sum(dep)-(final_easting-initial_easting)
    # print(er_dep,er_lat)
    corr_lat , correction_lat = T_correction_bowditch(leg_lengths,lat,er_lat) if typ=='bow' else T_correction_transit(lat,er_lat)
    corr_dep , correction_dep = T_correction_bowditch(leg_lengths,dep,er_dep) if typ=='bow' else T_correction_transit(dep,er_dep)

    # print(sum(corr_lat),sum(corr_dep))

    write_to_column(sheet, 18, 8, len(dep),correction_dep)
    write_to_column(sheet, 19, 8, len(lat),correction_lat)
    write_to_column(sheet, 20, 8, len(dep),corr_dep)
    write_to_column(sheet, 21, 8, len(lat),corr_lat)


    eastings = []
    northings = []
    eastings.append(initial_easting)
    northings.append(initial_northing)

    for i in range(n_minor+1):
        north = northings[i]+corr_lat[i]
        east = eastings[i]+corr_dep[i]
        eastings.append(east)
        northings.append(north)

    write_to_column(sheet, 22, 8, n_minor+2, eastings)
    write_to_column(sheet, 23, 8, n_minor+2, northings)


    al = []
    cb = []
    for i in range (n_minor+1):
        dE = eastings[i+1]-eastings[i]
        dN = northings[i+1]-northings[i]

        alength = math.sqrt(dE**2+dN**2)
        quad = math.atan(dE/dN)
        wcb = qbToWcb(dE,dN,quad*(180/math.pi))
        al.append(alength)
        cb.append(f'{wcb[0]}° {wcb[1]}\' {wcb[2]}\"')

    write_to_column(sheet, 24, 8, n_minor+1, al)
    write_to_column(sheet, 25, 8, n_minor+1, cb)

    try:
        wb.save(path)
    except Exception as e:
        Ui_MainWindow.failed('Unexpected error occured while saving file\nMake sure to close the excel file after adding data','file save failed')
        return 0

    
    Ui_MainWindow.state['is_calculated'] = 'T'
    Ui_MainWindow.state['north'] = northings
    Ui_MainWindow.state['east'] = eastings
    Ui_MainWindow.state['ang_err'] = bea_err
    Ui_MainWindow.state['lat_err'] = er_lat
    Ui_MainWindow.state['dep_err'] = er_dep
    Ui_MainWindow.state['perimeter'] = perimeter


    msg = QMessageBox()
    msg.setIcon(QMessageBox.Information)
    msg.setWindowTitle('Saved')
    msg.setText('Calculation Completed and Saved')
    x = msg.exec_()


    

class Ui_MainWindow(object):

    state = {
        'path' : '',
        'n' : 0,
        'tra_type' : 'closed',
        'type' : '',
        'is_calculated' : 'F',
        'north' : [],
        'east' : [],
        'ang_err' : [],
        'lat_err' : 0,
        'dep_err' : 0,
        'perimeter' : 0
    }

    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(545, 545)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("tb.ico"), QtGui.QIcon.Normal, QtGui.QIcon.On)
        MainWindow.setWindowIcon(icon)

        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setStyleSheet("background:rgb(239, 239, 239)")
        self.centralwidget.setObjectName("centralwidget")

        self.group1 = QtWidgets.QGroupBox(self.centralwidget)
        self.group1.setGeometry(QtCore.QRect(10, 10, 520, 130))
        self.group1.setObjectName("group1")

        self.label1 = QtWidgets.QLabel(self.group1)
        self.label1.setGeometry(QtCore.QRect(10, 20, 130, 30))
        font = QtGui.QFont()
        font.setPointSize(14)
        self.label1.setFont(font)
        self.label1.setObjectName("label1")

        self.spinBox1 = QtWidgets.QSpinBox(self.group1)
        self.spinBox1.setGeometry(QtCore.QRect(150, 20, 50, 30))
        self.spinBox1.setMouseTracking(True)
        self.spinBox1.setObjectName("spinBox1")

        self.line = QtWidgets.QFrame(self.centralwidget)
        self.line.setGeometry(QtCore.QRect(0, 140, 545, 16))
        self.line.setFrameShape(QtWidgets.QFrame.HLine)
        self.line.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line.setObjectName("line")

        self.create_btn = QtWidgets.QPushButton(self.group1)
        self.create_btn.setGeometry(QtCore.QRect(20, 80, 100, 30))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.create_btn.setFont(font)
        self.create_btn.setObjectName("create_btn")

        self.label_2 = QtWidgets.QLabel(self.group1)
        self.label_2.setGeometry(QtCore.QRect(140, 90, 321, 16))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")

        self.comboBox = QtWidgets.QComboBox(self.group1)
        self.comboBox.setGeometry(QtCore.QRect(370, 20, 145, 30))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.comboBox.setFont(font)
        self.comboBox.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.comboBox.setObjectName("comboBox")
        self.comboBox.addItem("")
        self.comboBox.addItem("")

        self.label_6 = QtWidgets.QLabel(self.group1)
        self.label_6.setGeometry(QtCore.QRect(230, 20, 130, 30))
        font = QtGui.QFont()
        font.setPointSize(14)
        self.label_6.setFont(font)
        self.label_6.setObjectName("label_6")

        self.groupBox = QtWidgets.QGroupBox(self.centralwidget)
        self.groupBox.setGeometry(QtCore.QRect(10, 160, 520, 210))
        self.groupBox.setObjectName("groupBox")

        self.label_4 = QtWidgets.QLabel(self.groupBox)
        self.label_4.setGeometry(QtCore.QRect(20, 20, 491, 16))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_4.setFont(font)
        self.label_4.setObjectName("label_4")

        self.label = QtWidgets.QLabel(self.groupBox)
        self.label.setGeometry(QtCore.QRect(20, 40, 141, 16))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label.setFont(font)
        self.label.setObjectName("label")

        self.calculate_btn = QtWidgets.QPushButton(self.groupBox)
        self.calculate_btn.setGeometry(QtCore.QRect(20, 140, 101, 30))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.calculate_btn.setFont(font)
        self.calculate_btn.setObjectName("calculate_btn")

        self.label_3 = QtWidgets.QLabel(self.groupBox)
        self.label_3.setGeometry(QtCore.QRect(20, 185, 380, 15))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_3.setFont(font)
        self.label_3.setObjectName("label_3")

        self.trans = QtWidgets.QRadioButton(self.groupBox)
        self.trans.setGeometry(QtCore.QRect(20, 110, 100, 20))
        self.trans.setChecked(True)
        self.trans.setObjectName("trans")

        self.bow = QtWidgets.QRadioButton(self.groupBox)
        self.bow.setGeometry(QtCore.QRect(140, 110, 110, 20))
        self.bow.setChecked(False)
        self.bow.setObjectName("bow")

        self.line_2 = QtWidgets.QFrame(self.centralwidget)
        self.line_2.setGeometry(QtCore.QRect(0, 370, 545, 16))
        self.line_2.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_2.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_2.setObjectName("line_2")

        self.label_5 = QtWidgets.QLabel(self.groupBox)
        self.label_5.setGeometry(QtCore.QRect(20, 80, 310, 25))
        font = QtGui.QFont()
        font.setPointSize(14)
        self.label_5.setFont(font)
        self.label_5.setObjectName("label_5")

        self.group3 = QtWidgets.QGroupBox(self.centralwidget)
        self.group3.setGeometry(QtCore.QRect(10, 390, 520, 100))
        self.group3.setObjectName("group3")



        self.show_btn = QtWidgets.QPushButton(self.group3)
        self.show_btn.setGeometry(QtCore.QRect(60, 30, 150, 30))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.show_btn.setFont(font)
        self.show_btn.setObjectName("show_btn")

        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 545, 21))
        self.menubar.setObjectName("menubar")
        self.menuAbout = QtWidgets.QMenu(self.menubar)
        self.menuAbout.setObjectName("menuAbout")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        self.actionCreator = QtWidgets.QAction(MainWindow)
        self.actionCreator.setObjectName("actionCreator")
        self.actionApp = QtWidgets.QAction(MainWindow)
        self.actionApp.setObjectName("actionApp")
        self.actionContact = QtWidgets.QAction(MainWindow)
        self.actionContact.setCheckable(False)
        self.actionContact.setObjectName("actionContact")
        self.menuAbout.addAction(self.actionCreator)
        self.menuAbout.addAction(self.actionApp)
        self.menuAbout.addAction(self.actionContact)
        self.menubar.addAction(self.menuAbout.menuAction())


        self.actionCreator.triggered.connect(self.Creator)
        self.actionApp.triggered.connect(self.App)
        self.actionContact.triggered.connect(self.Contact)

        self.create_btn.clicked.connect(self.create)
        self.calculate_btn.clicked.connect(self.calculate)
        self.show_btn.clicked.connect(self.show)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)


    def Creator(self):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setText('Author : Ritesh Ghimere\nProfession : Student @Thapathali Engineering Campus\n(batch : 2074/75 to 2078/79)')
        msg.setWindowTitle('Creator')
        x = msg.exec_()

    def App(self):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setText('App : Traverse Calculator\nVersion : 1.3.0.0\n\nUse only for Educational Purpose\n\n\nCopyright : RiteshGhimere © 2019')
        msg.setWindowTitle('App')
        x = msg.exec_()

    def Contact(self):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setText('Email : cool.rkroking@gmail.com\n\nFeel free to give Feedback or recomend any suggestions,problems')
        msg.setWindowTitle('Contact')
        x = msg.exec_()       
    
    def create(self):
        if(self.spinBox1.value()<3):
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setText("Number of Stations must be grater than 2")
            msg.setWindowTitle("Warning")
            x = msg.exec_()
            return 0

        self.state['tra_type'] = 'closed' if self.comboBox.currentIndex() == 0 else 'link'
        print(self.state['tra_type'])
        filename = QFileDialog.getSaveFileName(self.centralwidget,'Single File','d:\'','*.xlsx')
        path = filename[0]
        print(path)
        if(path == None):
            return 0
        self.state['path'] = path
        self.state['n'] = self.spinBox1.value()
        createTable(self.state['path'],self.state['n']) if self.state['tra_type'] == 'closed' else createMTable(self.state['path'],self.state['n']) 

    def failed(message,title):
        msg = QMessageBox()
        msg.setText(message)
        msg.setIcon(QMessageBox.Critical)
        msg.setWindowTitle(title)
        x = msg.exec_()
        return 0

    def calculate(self):
        if(self.trans.isChecked()):
            self.state['type'] = 'trans'
        if(self.bow.isChecked()):
            self.state['type'] = 'bow'
        path = self.state['path']
        typ = self.state['type']
        calculate_traverse(path,typ) if self.state['tra_type'] == 'closed' else calculate_Mtraverse(path,typ)

    def show(self):
        # check if calculation is completed
        if(self.state['is_calculated'] == 'F'):
            return 0

        north = []
        east = []
        # get calculated data from state object
        for i in range(len(self.state['north'])):
            north.append(self.state['north'][i])
            east.append(self.state['east'][i])

        y = []
        x = []
        
        # Reflection about X-axis
        for i in range(len(north)):
            north[i] = -north[i]

        # Determining scale for fitting the traverse in 500px height and 700px width
        scale1 = round(700/(max(east)-min(east)),4)
        scale2 = round(500/(max(north)-min(north)),4)
        scale = scale1 if scale1<scale2 else scale2
        # scaling all coordinates by factor acquired factor
        for i in range(len(north)):
            north[i] = north[i] * scale
            east[i] = east[i] * scale

        # translating coordinates onto screen area
        mn_north = min(north)
        mn_east = min(east)
        for i in range(len(north)):
            y.append(int(north[i] - mn_north + 75))
            x.append(int(east[i] - mn_east + 50))

        running = True
        win_size = [1150,600]
        Color_screen=[201, 169, 167]
        Color_line=(0,0,0)

        # initialize pygame
        pygame.init()
        scr = pygame.display.set_mode(win_size)
        pygame.display.set_caption('Traverse')
        scr.fill(Color_screen)
        pygame.display.flip()
        font_16 = pygame.font.Font('Muli-Regular.ttf', 16)
        font_28 = pygame.font.Font('Muli-Regular.ttf', 28)
        font_22 = pygame.font.Font('Muli-Regular.ttf', 22)
        font_uc = pygame.font.Font('Calibri.ttf', 22)
        # Drawing Traverse
        if(self.state['tra_type'] == 'closed'):
            for i in range(len(x)-1):
                pygame.time.delay(500)
                pygame.draw.line(scr,Color_line,(x[i],y[i]),(x[i+1],y[i+1]),2)
                pygame.draw.circle(scr,Color_line,(x[i],y[i]),5,1)
                pygame.draw.circle(scr,Color_line,(x[i],y[i]),10,1)
                text = font_16.render(f'M{i+1}', True, Color_line)
                textRect = text.get_rect()
                textRect.center = (20, 20)
                scr.blit(text,[x[i]-30,y[i]-30])
                pygame.display.flip()
        else:
            text = font_16.render(f'P', True, Color_line)
            textRect = text.get_rect()
            textRect.center = (20, 20)
            scr.blit(text,[x[0]-30,y[0]-30])
            pygame.draw.circle(scr,Color_line,(x[0],y[0]),5,1)
            pygame.draw.circle(scr,Color_line,(x[0],y[0]),10,1)
            pygame.display.flip()

            for i in range(1,len(x)-1):
                pygame.time.delay(500)
                pygame.draw.line(scr,Color_line,(x[i-1],y[i-1]),(x[i],y[i]),2)
                pygame.draw.circle(scr,Color_line,(x[i],y[i]),10,1)
                text = font_16.render(f'm{i}', True, Color_line)
                textRect = text.get_rect()
                textRect.center = (20, 20)
                scr.blit(text,[x[i]-30,y[i]-30])
                pygame.display.flip()

            l = len(x)
            pygame.time.delay(500)
            pygame.draw.line(scr,Color_line,(x[l-2],y[l-2]),(x[l-1],y[l-1]),2)   
            text = font_16.render(f'R', True, Color_line)
            textRect = text.get_rect()
            textRect.center = (20, 20)
            scr.blit(text,[x[l-1]-30,y[l-1]-30])
            pygame.draw.circle(scr,Color_line,(x[l-1],y[l-1]),5,1)
            pygame.draw.circle(scr,Color_line,(x[l-1],y[l-1]),10,1)
            pygame.display.flip()

        # Displaying Scale
        pygame.time.delay(500)
        text = font_28.render(f'Scale : {scale} pixels = 1 metre', True, Color_line)
        textRect = text.get_rect()
        textRect.center = (100, 50)
        scr.blit(text,[100,10])
        pygame.display.flip()
        pygame.time.delay(500)

        Deg = self.state['ang_err'][0]
        Min = abs(self.state['ang_err'][1])
        Sec = abs(self.state['ang_err'][2])
        text = font_uc.render(f'Angular Error = {Deg}° {Min}\' {Sec}\"', True, Color_line)
        textRect = text.get_rect()
        textRect.center = (100, 50)
        scr.blit(text,[800,300])
        pygame.display.flip()
        pygame.time.delay(500)
        
        ΣL = round(self.state['lat_err'],4)

        text = font_uc.render(f'δL = {ΣL}', True, Color_line)
        textRect = text.get_rect()
        textRect.center = (100, 50)
        scr.blit(text,[800,330])
        pygame.display.flip()
        pygame.time.delay(500)

        ΣD = round(self.state['dep_err'],4)

        text = font_uc.render(f'δD = {ΣD}', True, Color_line)
        textRect = text.get_rect()
        textRect.center = (100, 50)
        scr.blit(text,[800,360])
        pygame.display.flip()
        pygame.time.delay(500)

        tot_err = round(math.sqrt((self.state['lat_err']**2)+(self.state['dep_err']**2)),4)

        text = font_uc.render(f'Closing Error = {tot_err}', True, Color_line)
        textRect = text.get_rect()
        textRect.center = (100, 50)
        scr.blit(text,[800,390])
        pygame.display.flip()
        pygame.time.delay(500)

        rel_pre = round(self.state['perimeter'] / tot_err,4)

        text = font_uc.render(f'Relative Precision = 1 : {rel_pre}', True, Color_line)
        textRect = text.get_rect()
        textRect.center = (100, 50)
        scr.blit(text,[800,420])
        pygame.display.flip()

        path = self.state['path']
        path_a = path.split('.')[0]
        path_b = path_a + '.png'
        print(path)
        try:
            pygame.image.save(scr,path_b)
        except Exception as e:
            text = font_22.render(f'image save failed', True, Color_line)
            textRect = text.get_rect()
            textRect.center = (100, 50)
            scr.blit(text,[700,50])
            pygame.display.flip()
        else:
            text = font_22.render(f'image saved : {path_b}', True, Color_line)
            textRect = text.get_rect()
            textRect.center = (100, 50)
            scr.blit(text,[700,50])
            pygame.display.flip()
        while running:
            # looking for events
            for events in pygame.event.get():
                if events.type == QUIT:
                    running = False
                    pygame.quit()



    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Traverse_calc"))
        self.group1.setTitle(_translate("MainWindow", "Create Gales Table"))
        self.label1.setText(_translate("MainWindow", "No of Stations : "))
        self.create_btn.setText(_translate("MainWindow", "Create"))
        self.label_6.setText(_translate("MainWindow", "Traverse Type :"))
        self.comboBox.setItemText(0, _translate("MainWindow", "Closed Traverse"))
        self.comboBox.setItemText(1, _translate("MainWindow", "Link Traverse"))
        self.label_2.setText(_translate("MainWindow", "This will create an excel file with required gales table"))
        self.groupBox.setTitle(_translate("MainWindow", "Calculate"))
        self.label_4.setText(_translate("MainWindow", "open above created file, fill out all required data marked with red colour"))
        self.label.setText(_translate("MainWindow", "and click on calculate "))
        self.calculate_btn.setText(_translate("MainWindow", "Calculate"))
        self.label_3.setText(_translate("MainWindow", "the calculated data will be saved within same file created above"))
        self.trans.setText(_translate("MainWindow", "Transit Method"))
        self.bow.setText(_translate("MainWindow", "Bowditchs Method"))
        self.label_5.setText(_translate("MainWindow", "Correction Type"))
        self.group3.setTitle(_translate("MainWindow", "Display"))

        self.show_btn.setText(_translate("MainWindow", "Draw Traverse"))
        self.menuAbout.setTitle(_translate("MainWindow", "About"))
        self.actionCreator.setText(_translate("MainWindow", "Creator"))
        self.actionApp.setText(_translate("MainWindow", "App"))
        self.actionContact.setText(_translate("MainWindow", "Contact"))


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
